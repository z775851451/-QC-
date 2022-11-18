# %%
#coding:utf-8
# import _scproxy
import pymssql
import pandas as pd
import numpy as np
import openpyxl
import datetime
from openpyxl import load_workbook
import json  
import warnings


warnings.filterwarnings('ignore')

def sql_connect(server='192.168.0.15',user='zhongxin_zyanbo',password='ZhangYB_068',database='QC',sql=None):
    syntun_conn = pymssql.connect(server=server,
                            user=user,
                            password=password,
                            database=database)
    syntun_cursor = syntun_conn.cursor()

    syntun_cursor.execute(sql)
    s = syntun_cursor.fetchall()
    syntun_cursor.close()
    syntun_conn.close()
    return s

# %%
# x_df = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '客户产品名称')
# x_df_zy = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '平台自营判断')
# x_df_gzys = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '整体映射规则')
# x_df_gzys2 = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '独立映射规则')

x_df = pd.DataFrame(sql_connect(sql = 'select CAST ( 客户 AS nvarchar ( 500 ) ),CAST ( 数据库名 AS nvarchar ( 500 ) ),CAST ( 平台名称 AS nvarchar ( 500 ) ),CAST ( 店铺名称 AS nvarchar ( 500 ) ),CAST ( 判断制造商 AS nvarchar ( 500 ) ),CAST ( 判断品牌 AS nvarchar ( 500 ) ),CAST ( 判断子品牌 AS nvarchar ( 500 ) ),CAST ( 店铺类型 AS nvarchar ( 500 ) ),CAST ( 规则 AS nvarchar ( 500 ) ),CAST ( 判断月份 AS nvarchar ( 500 ) )  from 店铺名称判断店铺类型异常_对照表'),columns = ['客户','数据库名','平台名称','店铺名称','判断制造商','判断品牌','判断子品牌','店铺类型','规则','判断月份'])
x_df_zy = pd.DataFrame(sql_connect(sql = 'select CAST ( 店铺名称 AS nvarchar ( 500 ) ),CAST ( 店铺类型 AS nvarchar ( 500 ) ) from 店铺名称判断店铺类型异常_平台自营判断'),columns = ['店铺名称','店铺类型'])
x_df_gzys = pd.DataFrame(sql_connect(sql = 'select CAST ( 规则名称 AS nvarchar ( 500 ) ),CAST ( 备注 AS nvarchar ( 500 ) ) ,CAST ( 规则类型 AS nvarchar ( 500 ) )from 店铺名称判断店铺类型异常_整体映射规则'),columns = ['规则名称','备注','规则类型'])
x_df_gzys2 = pd.DataFrame(sql_connect(sql = 'select CAST ( 店铺名称 AS nvarchar ( 500 ) ),CAST ( 店铺类型 AS nvarchar ( 500 ) )  from 店铺名称判断店铺类型异常_独立映射规则'),columns = ['店铺名称','店铺类型'])

# x_df.数据库名.to_list()[0]
# %%
print('客户列表')
print(x_df['客户'].to_list())
# %%
#20221114
inp_date = input('仅运行当前日期之后(格式:202201):')
inp_ku = input('仅运行所选客户,逗号分隔(蒙牛,伊利(以模版文件客户名为准,运行全部 输入all)):').split(',')

if inp_ku[0] != 'all':
    x_df = x_df[x_df['客户'].isin(inp_ku)].reset_index()
else:
    pass

# %%
dict_all = {}
for i in range(len(x_df_gzys)):
    dict_all.update(json.loads(x_df_gzys['备注'][i]))
    

# %%
# Z = {
#     10:'其他',
#     1:'自营',
#     2:['品牌旗舰店','京东品牌旗舰店','天猫品牌旗舰店'],
#     3:'卖场型旗舰店',
#     4:'专卖店',
#     5:'专营店'
# }

z = {**{
    10:10,
    1:1,
    2:1,
    3:3,
    4:4,
    5:5
},**dict(zip(x_df_zy['店铺名称'],x_df_zy['店铺类型']))}

# %%

# %%
def sql_connect(server = '192.168.0.15',user = 'zhongxin_yanfa',password = 'Xin_yanfa',database = None,sql = None,no=0):
    syntun_conn = pymssql.connect(server=server,
                              user=user,
                              password=password,
                              database=database)
    syntun_cursor = syntun_conn.cursor()
    sql =  f"select distinct \
                            CAST ( {x_df.平台名称[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.店铺名称[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.判断制造商[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.判断品牌[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.判断子品牌[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.店铺类型[no]} AS nvarchar ( 500 ) )\
                            FROM {x_df.数据库名[no]} \
                            where {x_df.店铺类型[no]} is not null and {x_df.店铺类型[no]} != '海外购' and {x_df.判断品牌[no]} != '未知'\
                            and SUBSTRING(REPLACE({x_df.判断月份[no]}, '-', ''), 0, 7) >= {inp_date}" 
                                
    syntun_cursor.execute(sql)
    s = syntun_cursor.fetchall()
    syntun_cursor.close()
    syntun_conn.close()
    
    df = pd.DataFrame(s,columns = ['平台名称','店铺名称','制造商','品牌','子品牌','店铺类型'])
    
    return df

# %%
#相似度
xsd = 0.8

# %%
import difflib
 
def idf(s1, s2):
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()

idf('官方旗舰店 220ml','旗舰店官方 220ml')
# xsd = 0.7
# df[df.apply(lambda x: ('旗舰店'in(x.店铺名称) ) & ((x.制造商 in(x.店铺名称))|(x.品牌 in(x.店铺名称))|(idf(lambda x: x in'旗舰店',x.店铺名称,x.制造商) >= xsd)|(idf(lambda x: x in'旗舰店',x.店铺名称,x.品牌) >= xsd)|(idf(lambda x: x in'品牌旗舰店',x.店铺名称,x.制造商) >= xsd)|(idf(lambda x: x in'品牌旗舰店',x.店铺名称,x.品牌) >= xsd)) ,axis=1)]

# idf('N30°官方旗舰店','n30°')

# %%
dl_pp = x_df_gzys2[x_df_gzys2['店铺类型'] == '品牌旗舰店']['店铺名称'].to_list()
dl_mcx = x_df_gzys2[x_df_gzys2['店铺类型'] == '卖场型旗舰店']['店铺名称'].to_list()
dl_zm = x_df_gzys2[x_df_gzys2['店铺类型'] == '专卖店']['店铺名称'].to_list()
dl_zy = x_df_gzys2[x_df_gzys2['店铺类型'] == '专营']['店铺名称'].to_list()

# %%
class lg:
     def __init__(self):
          pass

     #自营 优先级1
     #1,若存在于[平台自营判断 - 规则表],不抛出
     def N1_lg(self,df): 
     
          T = df.apply(
                    #  lambda x: 1 if (('超市'in(x.店铺名称)) | ('自营'in(x.店铺名称)) )else 0
                    lambda x: x.店铺名称 if (x.店铺名称 in(x_df_zy['店铺名称'].to_list())) else 0
                    ,axis=1)
          return T

     #品牌旗舰店 优先级2
     #1,店铺名称 包含 ‘旗舰店’ 且
     #2,制造商 转为小写 存在于 店铺名称 转为小写 或
     #3,品牌 转为小写 存在于 店铺名称 转为小写 或
     #4,店铺名称 去除 ‘旗舰店’ (存在英文字符 转为小写) 与 制造商 转为小写 相似度 在0.7以上 或
     #5,店铺名称 去除 ‘旗舰店’ (存在英文字符 转为小写) 与 品牌\子品牌 转为小写 相似度 在0.7以上 或
     #6,店铺名称 去除 ‘品牌旗舰店’ (存在英文字符 转为小写) 与 制造商 转为小写 相似度 在0.7以上 或
     #7,店铺名称 去除 ‘品牌旗舰店’ (存在英文字符 转为小写) 与 品牌\子品牌 转为小写 相似度 在0.7以上
     
     def N2_lg(self,df):
          
          T = df.apply(
                    lambda x: 2 if ('旗舰店'in(x.店铺名称) ) & (
                              (x.制造商.lower() in(x.店铺名称.lower()))|
                              (x.品牌.lower() in(x.店铺名称.lower()))|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.制造商.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.子品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.制造商.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.子品牌.lower()) >= xsd)|
                              (x.店铺名称 in(dl_pp))
                         )  else 0
                    ,axis=1)
          return T

     #卖场型旗舰店 优先级3
     #1,店铺名称 包含 ‘旗舰店’ 且
     #2,制造商 不存在于 店铺名称 或
     #3,品牌 不存在于 店铺名称
     def N3_lg(self,df):
     
          T = df.apply(
                    lambda x:3 if ('旗舰店'in(x.店铺名称) ) & ((x.制造商 not in(x.店铺名称))|(x.品牌 not in(x.店铺名称)))|
                              (x.店铺名称 in(dl_mcx)) else 0
                    ,axis=1)
          return T

     #专卖店 优先级4
     #1,店铺名称 包含 ‘专卖’
     def N4_lg(self,df):
     
          T = df.apply(
                    lambda x:4 if ('专卖'in(x.店铺名称) )|
                              (x.店铺名称 in(dl_zm)) else 0
                    ,axis=1)
          return T

     #专营 优先级5
     #1,店铺名称 包含 ‘专营’
     def N5_lg(self,df):
     
          T = df.apply(
                    lambda x:5 if ('专营'in(x.店铺名称) )|
                              (x.店铺名称 in(dl_zy)) else 0
                    ,axis=1)
          return T

     def res(self,df):
          return df.assign(A = self.N1_lg(df.fillna('null')),B = self.N2_lg(df.fillna('null')),C = self.N3_lg(df.fillna('null')),D = self.N4_lg(df.fillna('null')),E=self.N5_lg(df.fillna('null')))



# %%
# pp = lg().res(sql_connect(no=15)).drop_duplicates(subset=['品牌'])['品牌']
# zpp = lg().res(sql_connect(no=15)).drop_duplicates(subset=['子品牌'])['子品牌']
# tt = pd.concat([pp,zpp]).drop_duplicates()

# %%
# x_df.数据库名[15]

# %%
# '特仑苏' in(tt.to_list())

# %%
# test_ = sql_connect(no=33)

# %%
# print(f'结果/品牌旗舰店判断规则表-结果_({inp_ku[0:2]}....xlsx')

# %%
res = []
for i in range(len(x_df)):
    print(x_df.数据库名[i])
    
    try:
        lg().res(sql_connect(no=i))
    except:
        continue
    
    df = lg().res(sql_connect(no=i))
    #
    pp = lg().res(sql_connect(no=i)).drop_duplicates(subset=['品牌'])['品牌']
    zpp = lg().res(sql_connect(no=i)).drop_duplicates(subset=['子品牌'])['子品牌']
    tt = pd.concat([pp,zpp]).drop_duplicates()
    
    if x_df.数据库名[i] == 'send_out.dbo.COKE_E_NEW':
        df = df[df['平台名称'] != '拼多多']
    Z = dict_all[x_df['规则'][i]]
    #0为自营,自营需要特殊处理
    df[['A','B','C','D','E']] = df[['A','B','C','D','E']].replace(0, 10)
    df['A'] = df['A'].map(z)
    #1,取A-E 最小值(意义:第一个能匹配上的类型)
    #2,转换匹配类型的字典
    #3,新增一列组合自营+其他所有
    df = df.assign(sum_ = df[['A','B','C','D','E']].min(axis=1).astype('int').astype('str'),
                lx = lambda x:x['sum_'].map(Z),
                程序_不一致 = lambda x:x.apply(lambda x :x.A if str(x.A) != '10' else x.lx,axis=1)
                # ,
                # 异常分类 = lambda x:x['sum_'].map(fl)
                )
    df = df[df[['店铺类型','程序_不一致']].apply( lambda x: x.店铺类型 not in x.程序_不一致,axis=1)][['平台名称','店铺名称','制造商','品牌','店铺类型','程序_不一致']]
    df.insert(0,'客户',x_df.客户[i])
    df.insert(1,'数据库名',x_df.数据库名[i])
    df = df.reset_index(drop=True)
    #
    #数据库是品牌旗舰店
    
    if x_df_gzys[(x_df_gzys['规则名称'] == x_df['规则'][i])]['规则类型'].to_list()[0] != 0:
        aa = df[(df['店铺类型'].isin(Z['2']))&(df['店铺名称'].str.contains('旗舰店'))]
        ppdc = [j if str(i).lower() in (str(j).lower()) else '0' for i in tt for j in aa['店铺名称']]
        aa['店铺名称'][aa['店铺名称'].isin(list(set(ppdc)))]
        
        lsmc = df.iloc[aa['店铺名称'][aa['店铺名称'].isin(list(set(ppdc)))].index,:]['店铺名称'].to_list()
        df = df.assign(异常分类 =  df.apply(lambda x: '品牌异常' if (x['店铺名称'] in(lsmc)) & (x['店铺类型']in(Z['2'])) else '类型异常',axis=1))
        df = df.assign(程序判定 =  df.apply(lambda x: Z['2'] if x.异常分类 == '品牌异常' else x.程序_不一致,axis=1))
        
        # df = df.assign(异常分类 =  df['店铺名称'].map(lambda x: '品牌异常' if x in(lsmc) else '类型异常'))
    else:
        df = df.assign(异常分类 = '类型异常')
        df = df.assign(程序判定 = df['程序_不一致'])
    
        
    res.append(df)
try:
    df_ = pd.concat(res)


    df_ = df_.drop_duplicates(subset=['客户','数据库名','平台名称','店铺名称','制造商','品牌','店铺类型'])[['客户','数据库名','平台名称','店铺名称','制造商','品牌','店铺类型','异常分类','程序判定']]
    # df_ = df_.assign(程序判定 = df_['店铺类型'] == '')

    sl = pd.DataFrame(list(df_['数据库名'].value_counts().to_dict().items()),
                    columns=['数据库名称', '抛出数量'])

    # import openpyxl
    # from openpyxl import load_workbook

    # bsgg_workbook = load_workbook(r'模版/品牌旗舰店判断规则表.xlsx')
    # bsgg_writer = pd.ExcelWriter(r'模版/品牌旗舰店判断规则表.xlsx',
    #                         engine='openpyxl')
    # bsgg_writer.book= bsgg_workbook
    # #防止模版损坏先保存一个
    # bsgg_workbook.save(r'模版/品牌旗舰店判断规则表.xlsx')



    # df_.to_excel(bsgg_writer, sheet_name='抛出',na_rep='',index=False)
    # sl.to_excel(bsgg_writer, sheet_name='抛出数量预览',na_rep='',index=False)

    # bsgg_workbook.save(f'结果/品牌旗舰店判断规则表-结果_{inp_ku[0:2]}等{len(inp_ku)}个库.xlsx')
    # bsgg_workbook.close()
    
    import openpyxl
    from openpyxl import load_workbook
    with pd.ExcelWriter(f'结果/品牌旗舰店判断规则表-结果_{inp_ku[0:2]}等{len(inp_ku)}个库.xlsx') as mn_writer:
        x_df.to_excel(mn_writer,sheet_name='客户',na_rep='',index=False)
        x_df_zy.to_excel(mn_writer,sheet_name='平台自营判断',na_rep='',index=False)
        x_df_gzys.to_excel(mn_writer,sheet_name='整体映射规则',na_rep='',index=False)
        x_df_gzys2.to_excel(mn_writer,sheet_name='独立映射规则',na_rep='',index=False)
        df_.to_excel(mn_writer,sheet_name='抛出',na_rep='',index=False)
        sl.to_excel(mn_writer,sheet_name='抛出数量预览',na_rep='',index=False)
        

except:
    print('当前运行的内容输出为空,已中断此次运行')
# %%
