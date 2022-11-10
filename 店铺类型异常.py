import _scproxy
import sys
import os
import pymssql
import pandas as pd
import numpy as np
import openpyxl
import datetime
from openpyxl import load_workbook
import json  

import os
def mkdir(path):
        folder = os.path.exists(path)
        if not folder:    
                os.makedirs(path)            #makedirs 创建文件时如果路径不存在会创建这个路径
                print('检测无 [模版] 文件夹,程序将自动创建,请将模版( 品牌旗舰店判断规则表.xlsx )放置到此处')#判断是否存在文件夹如果不存在则创建为文件夹
                input('放置后确认将运行')
        else:
                # print('正在存放至 [模版] 📁')
                pass
mkdir('模版')


x_df = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '客户产品名称')
x_df_zy = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '平台自营判断')
x_df_gzys = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '整体映射规则')
x_df_gzys2 = pd.read_excel('模版/品牌旗舰店判断规则表.xlsx',sheet_name = '独立映射规则')
# x_df.数据库名.to_list()[0]


dict_all = {}
for i in range(len(x_df_gzys)):
    dict_all.update(json.loads(x_df_gzys['备注'][i]))
    


# Z = {
#     10:'其他',
#     1:'自营',
#     2:['品牌旗舰店','京东品牌旗舰店','天猫品牌旗舰店'],
#     3:'卖场型旗舰店',
#     4:'专卖店',
#     5:'专营店'
# }

z = {**{
    10:10,
    1:1,
    2:1,
    3:3,
    4:4,
    5:5
},**dict(zip(x_df_zy['店铺名称'],x_df_zy['店铺类型']))}





def sql_connect(server = '192.168.0.15',user = 'zhongxin_yanfa',password = 'Xin_yanfa',database = None,sql = None,no=0):
    syntun_conn = pymssql.connect(server=server,
                              user=user,
                              password=password,
                              database=database)
    syntun_cursor = syntun_conn.cursor()
    sql =  f"select distinct \
                            CAST ( {x_df.平台名称[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.店铺名称[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.判断制造商[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.判断品牌[no]} AS nvarchar ( 500 ) ),\
                            CAST ( {x_df.店铺类型[no]} AS nvarchar ( 500 ) )\
                            FROM {x_df.数据库名[no]} \
                            where {x_df.店铺类型[no]} is not null and {x_df.店铺类型[no]} != '海外购'" 
                                
    syntun_cursor.execute(sql)
    s = syntun_cursor.fetchall()
    syntun_cursor.close()
    syntun_conn.close()
    
    df = pd.DataFrame(s,columns = ['平台名称','店铺名称','制造商','品牌','店铺类型'])
    
    return df


#相似度
xsd = 0.8


import difflib
 
def idf(s1, s2):
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()

# idf('a','av')
# xsd = 0.7
# df[df.apply(lambda x: ('旗舰店'in(x.店铺名称) ) & ((x.制造商 in(x.店铺名称))|(x.品牌 in(x.店铺名称))|(idf(lambda x: x in'旗舰店',x.店铺名称,x.制造商) >= xsd)|(idf(lambda x: x in'旗舰店',x.店铺名称,x.品牌) >= xsd)|(idf(lambda x: x in'品牌旗舰店',x.店铺名称,x.制造商) >= xsd)|(idf(lambda x: x in'品牌旗舰店',x.店铺名称,x.品牌) >= xsd)) ,axis=1)]

# idf('N30°官方旗舰店','n30°')


dl_pp = x_df_gzys2[x_df_gzys2['店铺类型'] == '品牌旗舰店']['店铺名称'].to_list()
dl_mcx = x_df_gzys2[x_df_gzys2['店铺类型'] == '卖场型旗舰店']['店铺名称'].to_list()
dl_zm = x_df_gzys2[x_df_gzys2['店铺类型'] == '专卖店']['店铺名称'].to_list()
dl_zy = x_df_gzys2[x_df_gzys2['店铺类型'] == '专营']['店铺名称'].to_list()



class lg:
     def __init__(self):
          pass
     
     #自营 优先级1
     #1,若存在于[平台自营判断 - 规则表],不抛出
     def N1_lg(self,df): 
     
          T = df.apply(
                    #  lambda x: 1 if (('超市'in(x.店铺名称)) | ('自营'in(x.店铺名称)) )else 0
                    lambda x: x.店铺名称 if (x.店铺名称 in(x_df_zy['店铺名称'].to_list())) else 0
                    ,axis=1)
          return T

     #品牌旗舰店 优先级2
     #1,店铺名称 包含 ‘旗舰店’ 且
     #2,制造商 转为小写 存在于 店铺名称 转为小写 或
     #3,品牌 转为小写 存在于 店铺名称 转为小写 或
     #4,店铺名称 去除 ‘旗舰店’ (存在英文字符 转为小写) 与 制造商 转为小写 相似度 在0.7以上 或
     #5,店铺名称 去除 ‘旗舰店’ (存在英文字符 转为小写) 与 品牌\子品牌 转为小写 相似度 在0.7以上 或
     #6,店铺名称 去除 ‘品牌旗舰店’ (存在英文字符 转为小写) 与 制造商 转为小写 相似度 在0.7以上 或
     #7,店铺名称 去除 ‘品牌旗舰店’ (存在英文字符 转为小写) 与 品牌\子品牌 转为小写 相似度 在0.7以上
     
     def N2_lg(self,df):
          
          T = df.apply(
                    lambda x: 2 if ('旗舰店'in(x.店铺名称) ) & (
                              (x.制造商.lower() in(x.店铺名称.lower()))|
                              (x.品牌.lower() in(x.店铺名称.lower()))|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.制造商.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('旗舰店', '').lower(),x.子品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.制造商.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.品牌.lower()) >= xsd)|
                              (idf(x.店铺名称.replace('品牌旗舰店', '').lower(),x.子品牌.lower()) >= xsd)|
                              (x.店铺名称 in(dl_pp))
                         )  else 0
                    ,axis=1)
          return T

     #卖场型旗舰店 优先级3
     #1,店铺名称 包含 ‘旗舰店’ 且
     #2,制造商 不存在于 店铺名称 或
     #3,品牌 不存在于 店铺名称
     def N3_lg(self,df):
     
          T = df.apply(
                    lambda x:3 if ('旗舰店'in(x.店铺名称) ) & ((x.制造商 not in(x.店铺名称))|(x.品牌 not in(x.店铺名称)))|
                              (x.店铺名称 in(dl_mcx)) else 0
                    ,axis=1)
          return T

     #专卖店 优先级4
     #1,店铺名称 包含 ‘专卖’
     def N4_lg(self,df):
     
          T = df.apply(
                    lambda x:4 if ('专卖'in(x.店铺名称) )|
                              (x.店铺名称 in(dl_zm)) else 0
                    ,axis=1)
          return T

     #专营 优先级5
     #1,店铺名称 包含 ‘专营’
     def N5_lg(self,df):
     
          T = df.apply(
                    lambda x:5 if ('专营'in(x.店铺名称) )|
                              (x.店铺名称 in(dl_zy)) else 0
                    ,axis=1)
          return T

     def res(self,df):
          return df.assign(A = self.N1_lg(df.fillna('null')),B = self.N2_lg(df.fillna('null')),C = self.N3_lg(df.fillna('null')),D = self.N4_lg(df.fillna('null')),E=self.N5_lg(df.fillna('null')))

from tqdm import tqdm
from time import sleep

res = []
for i in range(len(x_df)):
    print(x_df.数据库名[i])
    df = lg().res(sql_connect(no=i))
    #
    pp = lg().res(sql_connect(no=i)).drop_duplicates(subset=['品牌'])['品牌']
    zpp = lg().res(sql_connect(no=i)).drop_duplicates(subset=['子品牌'])['子品牌']
    tt = pd.concat([pp,zpp]).drop_duplicates()
    
    if x_df.数据库名[i] == 'send_out.dbo.COKE_E_NEW':
        df = df[df['平台名称'] != '拼多多']
    Z = dict_all[x_df['规则'][i]]
    #0为自营,自营需要特殊处理
    df[['A','B','C','D','E']] = df[['A','B','C','D','E']].replace(0, 10)
    df['A'] = df['A'].map(z)
    #1,取A-E 最小值(意义:第一个能匹配上的类型)
    #2,转换匹配类型的字典
    #3,新增一列组合自营+其他所有
    df = df.assign(sum_ = df[['A','B','C','D','E']].min(axis=1).astype('int').astype('str'),
                lx = lambda x:x['sum_'].map(Z),
                程序_不一致 = lambda x:x.apply(lambda x :x.A if str(x.A) != '10' else x.lx,axis=1)
                # ,
                # 异常分类 = lambda x:x['sum_'].map(fl)
                )
    df = df[df[['店铺类型','程序_不一致']].apply( lambda x: x.店铺类型 not in x.程序_不一致,axis=1)][['平台名称','店铺名称','制造商','品牌','店铺类型','程序_不一致']]
    df.insert(0,'客户',x_df.客户[i])
    df.insert(1,'数据库名',x_df.数据库名[i])
    df = df.reset_index(drop=True)
    #
    #数据库是品牌旗舰店
    
    if x_df_gzys[(x_df_gzys['规则名称'] == x_df['规则'][i])]['规则类型'].to_list()[0] != 0:
        aa = df[(df['店铺类型'].isin(Z['2']))&(df['店铺名称'].str.contains('旗舰店'))]
        ppdc = [j if str(i).lower() in (str(j).lower()) else '0' for i in tt for j in aa['店铺名称']]
        aa['店铺名称'][aa['店铺名称'].isin(list(set(ppdc)))]
        
        lsmc = df.iloc[aa['店铺名称'][aa['店铺名称'].isin(list(set(ppdc)))].index,:]['店铺名称'].to_list()
        df = df.assign(异常分类 =  df.apply(lambda x: '品牌异常' if (x['店铺名称'] in(lsmc)) & (x['店铺类型']in(Z['2'])) else '类型异常',axis=1))
        df = df.assign(程序判定 =  df.apply(lambda x: Z['2'] if x.异常分类 == '品牌异常' else x.程序_不一致,axis=1))
        
        # df = df.assign(异常分类 =  df['店铺名称'].map(lambda x: '品牌异常' if x in(lsmc) else '类型异常'))
    else:
        df = df.assign(异常分类 = '类型异常')
        df = df.assign(程序判定 = df['程序_不一致'])
    
        
    res.append(df)
df_ = pd.concat(res)


# df.apply(lambda x: '品牌异常' if (x['店铺名称'] in(lsmc)) & (x['店铺类型']in(Z['2'])) else '类型异常',axis=1)


sl = pd.DataFrame(list(df_['数据库名'].value_counts().to_dict().items()),
                   columns=['数据库名称', '抛出数量'])


import openpyxl
from openpyxl import load_workbook



def mkdir(path):
        folder = os.path.exists(path)
        if not folder:    
                os.makedirs(path)            #makedirs 创建文件时如果路径不存在会创建这个路径
                print('检测无结果文件夹,程序将自动创建 📁')#判断是否存在文件夹如果不存在则创建为文件夹
        else:
                print('正在存放至 [结果] 📁')
                pass
mkdir('结果')

bsgg_workbook = load_workbook('模版/品牌旗舰店判断规则表.xlsx')
bsgg_writer = pd.ExcelWriter('模版/品牌旗舰店判断规则表.xlsx',
                        engine='openpyxl')
bsgg_writer.book= bsgg_workbook
#防止模版损坏先保存一个
bsgg_workbook.save('模版/品牌旗舰店判断规则表.xlsx')


df_.to_excel(bsgg_writer, sheet_name='抛出',na_rep='',index=False)
sl.to_excel(bsgg_writer, sheet_name='抛出数量预览',na_rep='',index=False)

bsgg_workbook.save('结果/品牌旗舰店判断规则表.xlsx')
bsgg_workbook.close()